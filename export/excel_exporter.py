"""
export/excel_exporter.py — Excel(.xlsx) 내보내기

openpyxl로 다크 스타일 적용 Excel 파일 생성.
"""
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from db.models import Product


# 컬럼 정의: (헤더명, dict 키, 너비)
COLUMNS = [
    ("플랫폼",    "platform",        12),
    ("상품명",    "name",            45),
    ("브랜드",    "brand",           18),
    ("판매가",    "price",           12),
    ("정가",      "original_price",  12),
    ("칼로리",   "calories",         10),
    ("단백질(g)", "protein",         12),
    ("탄수화물(g)","carbs",          12),
    ("지방(g)",   "fat",             10),
    ("제조사",   "manufacturer",     18),
    ("원산지",   "origin",           14),
    ("리뷰수",   "review_count",     10),
    ("평점",     "rating",            8),
    ("수집일시", "collected_at",     18),
    ("상태",     "status",            8),
    ("URL",      "url",              50),
]


def export_excel(products: List[Product], path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품 리서치"

    # ── 헤더 스타일
    header_fill = PatternFill("solid", fgColor="2a2a3d")
    header_font = Font(name="Malgun Gothic", bold=True, color="E2E8F0", size=11)
    border_side = Side(style="thin", color="3d3d5c")
    cell_border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )

    # ── 헤더 행
    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── 데이터 행
    even_fill = PatternFill("solid", fgColor="252538")
    odd_fill = PatternFill("solid", fgColor="1e1e2e")
    data_font = Font(name="Malgun Gothic", color="E2E8F0", size=10)
    error_font = Font(name="Malgun Gothic", color="F87171", size=10)
    success_font = Font(name="Malgun Gothic", color="4ADE80", size=10)

    platform_colors = {
        "쿠팡": "E63312",
        "스마트스토어": "03C75A",
        "네이버쇼핑": "1EC800",
    }

    for row_idx, product in enumerate(products, 2):
        data = product.to_dict()
        is_even = (row_idx % 2 == 0)
        row_fill = even_fill if is_even else odd_fill

        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            raw_val = data.get(key, "")
            # 숫자형으로 저장할 컬럼
            if key in ("price", "original_price", "review_count"):
                try:
                    val = int(raw_val.replace(",", "")) if raw_val else None
                except (ValueError, AttributeError):
                    val = raw_val
            elif key in ("calories", "protein", "carbs", "fat", "rating"):
                try:
                    val = float(raw_val) if raw_val else None
                except (ValueError, AttributeError):
                    val = raw_val
            else:
                val = raw_val

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = cell_border

            # 폰트 결정
            if key == "status":
                if product.status == "success":
                    cell.font = success_font
                    cell.value = "✓ 성공"
                else:
                    cell.font = error_font
                    cell.value = "✗ 실패"
            elif product.status == "error":
                cell.font = error_font
            elif key == "platform":
                pcolor = platform_colors.get(product.platform, "7C6AF7")
                cell.font = Font(name="Malgun Gothic", bold=True, color=pcolor, size=10)
            else:
                cell.font = data_font

            # 정렬
            if key in ("price", "original_price", "calories", "protein",
                       "carbs", "fat", "review_count", "rating"):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif key == "url":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[row_idx].height = 22

    # ── 자동 필터
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(products) + 1}"

    # ── 메타 시트
    ws_meta = wb.create_sheet("정보")
    ws_meta["A1"] = "생성일시"
    ws_meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_meta["A2"] = "총 상품 수"
    ws_meta["B2"] = len(products)
    ws_meta["A3"] = "성공"
    ws_meta["B3"] = sum(1 for p in products if p.status == "success")
    ws_meta["A4"] = "실패"
    ws_meta["B4"] = sum(1 for p in products if p.status == "error")

    wb.save(path)
